import time
import pandas as pd
import os
import sys
from openpyxl import load_workbook
import numpy as np
import win32com.client as win32
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium import webdriver

chrome_options = webdriver.chrome.options.Options()
#chrome_options.add_argument("--headless")
chrome_options.add_argument('--disable-gpu')
# Starts driver
driver = webdriver.Chrome("C:\chromedriver\chromedriver.exe",options=chrome_options)
# driver = webdriver.PhantomJS()
# Gets page
driver.get("https://extranet.btgpactual.com/")
# Sets wait
wait = WebDriverWait(driver, 300)

time.sleep(2)
login = "tiago.nunes"
# sends login
driver.find_element_by_id("txtLogin").send_keys(login)
# Goes to Keyboard
driver.find_element_by_xpath("//*[@id='btnValidarLogin']").click()
time.sleep(2)
# Tries to bypass keyboard
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[11]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[2]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[11]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[4]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[54]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[18]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[21]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[27]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[31]").click()
driver.find_element_by_xpath("//*[@id='contentVirtualKeyboard']/div/div/div[22]").click()
#validates
driver.find_element_by_xpath("//*[@id='btnValidate']/span").click()
time.sleep(2)

driver.find_element_by_xpath("//*[@id='menu']/li[2]/a").click()
time.sleep(2)

driver.find_element_by_xpath("//*[@id='liCotas']/a/span").click()
time.sleep(1)

select = Select(driver.find_element_by_xpath("//*[@id='ddlFundos']"))

#565123=BALAN 550607=GRAPH
select.select_by_value("550607")

#driver.find_element_by_xpath("//*[@id='ddlIndexadores_chosen']/ul/li/input").send_keys("Cdi")
#time.sleep(1)

driver.find_element_by_xpath("//*[@id='ddlIndexadores_chosen']").click()
time.sleep(1)

driver.find_element_by_xpath("//*[@id='ddlIndexadores_chosen']/div/ul/li[6]").click()
time.sleep(1)

#Xml IBOV //*[@id="ddlIndexadores_chosen"]/div/ul/li[6]
#Xml First Date //*[@id="txtInicio"]
#Xml Second Date //*[@id="txtFim"]

driver.find_element_by_xpath("//*[@id='linkbtconsultar']/a").click()

result_file = os.path.join(r"C:\Downloads","Results.xls")
if os.path.exists(result_file):
    os.remove(result_file)
time.sleep(2)

driver.find_element_by_xpath("//*[@id='tblAtivoCarteira_wrapper']/div[2]/div[3]/div/div[2]/a[1]").click()
time.sleep(3)

driver.close()

print("Arquivo baixado")

result_file = os.path.join(r"C:\Downloads","Results.xls")
new_file = os.path.join(r"C:\Downloads","Graphdiario.xls")

if os.path.exists(new_file):
    os.remove(new_file)
if os.path.exists(r"C:\Downloads\Graphdiario.xlsx"):
    os.remove(r"C:\Downloads\Graphdiario.xlsx")

os.rename(result_file, new_file)

print("Arquivo renomeado")

#////////////////////////////////////////////////////////////////////////////////////////

fname = r"C:\Downloads\Graphdiario.xls"
excel = win32.Dispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)

wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wb.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()

print("Arquivo convertido para xlsx")

#/////////////////////////////////////////////////////////////////////////////////////////

try:
    new = pd.read_excel(r'C:\Downloads\Graphdiario.xlsx', 'Graphdiario', index_col=None, na_values=['NA'], date_format="YYYY-MM-DD")

except IOError:
    print("Excel novo não encontrado")

print("Essas são as informações que serão passadas para a lâmina")
print(new)

a=''

while a != 'y' and 'n':
    a=input("Você quer continuar? (y/n) ")

    if a=='n':
        sys.exit()

def append_df_to_excel(filename, df, sheet_name='Historico', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    writer = pd.ExcelWriter(filename, engine='openpyxl', date_format="YYYY-MM-DD", mode='a')


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


append_df_to_excel(r"\\192.168.1.5\lftm_asset\GESTAO\FUNDOS ABERTOS\LIFETIME GRAPHENE\Lâmina\Lamina_Graphene para acompanhamento\Acompanhamento_Graphene.xlsx",new,truncate_sheet=False,header=None)




import pandas as pd
from openpyxl import load_workbook
import numpy as np

try:
    new = pd.read_excel('C:\Downloads\Balandiario.xlsx', 'Balandiario', index_col=None, na_values=['NA'], date_format="YYYY-MM-DD")

except IOError:
    print("Excel novo não encontrado")

try:
    fund = pd.read_excel('Lamina_Balanced.xlsx', 'Historico', index_col=None, na_values=['NA'])

    #r'\\192.168.1.5\lftm_asset\GESTAO\FUNDOS ABERTOS\LIFETIME BALANCED\Lâminas\Lamina_Balanced.xlsx'
except IOError:
    print("Excel lâmina não encontrado")

new['Data de referência'] = new["Data de referência"].dt.strftime('%d/%m/%Y')
new['Data de referência'] = pd.to_datetime(new["Data de referência"])

print(new)

first_empty_row = np.where(fund.iloc[:, 0].isnull().values == True)

    #print(first_empty_row)

fer=int(first_empty_row[0][:1])


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
     # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', date_format="YYYY-MM-DD", mode='a')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
append_df_to_excel("Lamina_Balanced.xlsx",new,sheet_name="Historico",startrow=fer,truncate_sheet=False,header=None)